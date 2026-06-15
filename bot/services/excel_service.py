import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bot.database.connection import get_pool


def _header_style(ws, row, columns):
    fill = PatternFill("solid", fgColor="2E5FA3")
    font = Font(bold=True, color="FFFFFF")
    for col, val in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _border_all(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin")
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in r:
            cell.border = b


async def generate_daily_excel(shop_id: int, report_date: date) -> bytes:
    pool = await get_pool()
    async with pool.acquire() as conn:
        shop = await conn.fetchrow("SELECT name FROM shops WHERE id=$1", shop_id)
        sales = await conn.fetch(
            """SELECT s.id, s.date, s.total_som, s.paid_som,
                      COALESCE(c.name,'Naqd') as customer,
                      (SELECT COALESCE(SUM(si.qty*si.cost_price_som),0) FROM sale_items si WHERE si.sale_id=s.id) as cost
               FROM sales s LEFT JOIN customers c ON c.id=s.customer_id
               WHERE s.shop_id=$1 AND s.date=$2 ORDER BY s.created_at""",
            shop_id, report_date
        )
        sale_items = await conn.fetch(
            """SELECT si.*, p.name as product_name, s.date
               FROM sale_items si
               JOIN sales s ON s.id=si.sale_id
               JOIN products p ON p.id=si.product_id
               WHERE s.shop_id=$1 AND s.date=$2""",
            shop_id, report_date
        )
        expenses = await conn.fetch(
            "SELECT category, amount_som, amount_usd, description FROM expenses WHERE shop_id=$1 AND date=$2",
            shop_id, report_date
        )
        purchases = await conn.fetch(
            """SELECT p.total_som, p.paid_som, COALESCE(s.name,'Taminotchisiz') as supplier
               FROM purchases p LEFT JOIN suppliers s ON s.id=p.supplier_id
               WHERE p.shop_id=$1 AND p.date=$2""",
            shop_id, report_date
        )

    wb = Workbook()

    # ─── SOTUV VARAGI ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sotuvlar"
    ws1["A1"] = f"{shop['name']} — Sotuvlar {report_date.strftime('%d.%m.%Y')}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:F1")

    _header_style(ws1, 3, ["#", "Mijoz", "Jami (so'm)", "To'landi", "Qarz", "Foyda"])
    total_s, total_p, total_profit = 0, 0, 0
    for i, s in enumerate(sales, 1):
        debt = float(s["total_som"]) - float(s["paid_som"])
        profit = float(s["total_som"]) - float(s["cost"])
        ws1.append([i, s["customer"], float(s["total_som"]), float(s["paid_som"]), debt, profit])
        total_s += float(s["total_som"])
        total_p += float(s["paid_som"])
        total_profit += profit
    ws1.append(["", "JAMI", total_s, total_p, total_s - total_p, total_profit])
    ws1.cell(ws1.max_row, 1).font = Font(bold=True)
    _border_all(ws1, 3, ws1.max_row, 1, 6)

    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 18

    # ─── MAHSULOTLAR VARAGI ────────────────────────────────────────────
    ws2 = wb.create_sheet("Mahsulotlar")
    ws2["A1"] = f"Sotuv mahsulotlari — {report_date.strftime('%d.%m.%Y')}"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.merge_cells("A1:F1")
    _header_style(ws2, 3, ["Mahsulot", "Miqdor", "Narx (so'm)", "Chegirma", "Jami", "Tannarx"])
    for si in sale_items:
        total = float(si["qty"]) * float(si["sell_price_som"]) - float(si["discount_som"])
        ws2.append([
            si["product_name"], float(si["qty"]), float(si["sell_price_som"]),
            float(si["discount_som"]), total, float(si["cost_price_som"]) * float(si["qty"])
        ])
    _border_all(ws2, 3, ws2.max_row, 1, 6)
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # ─── XARAJATLAR VARAGI ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Xarajatlar")
    ws3["A1"] = f"Xarajatlar — {report_date.strftime('%d.%m.%Y')}"
    ws3["A1"].font = Font(bold=True, size=13)
    ws3.merge_cells("A1:D1")
    _header_style(ws3, 3, ["Tur", "Miqdor (so'm)", "Dollar ($)", "Izoh"])
    total_e = 0
    for e in expenses:
        ws3.append([e["category"], float(e["amount_som"]), float(e["amount_usd"]), e["description"] or ""])
        total_e += float(e["amount_som"])
    ws3.append(["JAMI", total_e, "", ""])
    _border_all(ws3, 3, ws3.max_row, 1, 4)
    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    # ─── KIRIMLAR VARAGI ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Kirimlar")
    ws4["A1"] = f"Kirimlar — {report_date.strftime('%d.%m.%Y')}"
    ws4["A1"].font = Font(bold=True, size=13)
    ws4.merge_cells("A1:C1")
    _header_style(ws4, 3, ["Taminotchi", "Jami (so'm)", "To'landi"])
    for p in purchases:
        ws4.append([p["supplier"], float(p["total_som"]), float(p["paid_som"])])
    _border_all(ws4, 3, ws4.max_row, 1, 3)
    for col in range(1, 4):
        ws4.column_dimensions[get_column_letter(col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def generate_monthly_excel(shop_id: int, year: int, month: int) -> bytes:
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, last_day)
    pool = await get_pool()
    async with pool.acquire() as conn:
        shop = await conn.fetchrow("SELECT name FROM shops WHERE id=$1", shop_id)
        daily = await conn.fetch(
            """SELECT
               s.date,
               COALESCE(SUM(s.total_som),0) as sales,
               COALESCE(SUM(si.qty*si.cost_price_som),0) as cost,
               (SELECT COALESCE(SUM(e.amount_som),0) FROM expenses e WHERE e.shop_id=$1 AND e.date=s.date) as exp
               FROM sales s
               JOIN sale_items si ON si.sale_id=s.id
               WHERE s.shop_id=$1 AND s.date BETWEEN $2 AND $3
               GROUP BY s.date ORDER BY s.date""",
            shop_id, start, end
        )
        top_products = await conn.fetch(
            """SELECT p.name, SUM(si.qty) as qty, SUM(si.qty*si.sell_price_som) as revenue,
                      SUM(si.qty*si.cost_price_som) as cost
               FROM sale_items si JOIN sales s ON s.id=si.sale_id JOIN products p ON p.id=si.product_id
               WHERE s.shop_id=$1 AND s.date BETWEEN $2 AND $3
               GROUP BY p.name ORDER BY revenue DESC LIMIT 20""",
            shop_id, start, end
        )

    MONTHS = ["Yanvar","Fevral","Mart","Aprel","May","Iyun","Iyul","Avgust","Sentabr","Oktabr","Noyabr","Dekabr"]
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Kunlik"
    ws1["A1"] = f"{shop['name']} — {MONTHS[month-1]} {year}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:F1")
    _header_style(ws1, 3, ["Sana", "Sotuv (so'm)", "Tannarx", "Yalpi foyda", "Xarajat", "Sof foyda"])
    totals = [0]*5
    for d in daily:
        sales = float(d["sales"])
        cost = float(d["cost"])
        exp = float(d["exp"])
        gross = sales - cost
        net = gross - exp
        ws1.append([str(d["date"]), sales, cost, gross, exp, net])
        for i, v in enumerate([sales, cost, gross, exp, net]):
            totals[i] += v
    ws1.append(["JAMI"] + totals)
    ws1.cell(ws1.max_row, 1).font = Font(bold=True)
    _border_all(ws1, 3, ws1.max_row, 1, 6)
    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 16

    ws2 = wb.create_sheet("Mahsulotlar")
    ws2["A1"] = f"TOP mahsulotlar — {MONTHS[month-1]} {year}"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.merge_cells("A1:D1")
    _header_style(ws2, 3, ["Mahsulot", "Miqdor", "Sotuv (so'm)", "Foyda (so'm)"])
    for p in top_products:
        ws2.append([p["name"], float(p["qty"]), float(p["revenue"]),
                    float(p["revenue"]) - float(p["cost"])])
    _border_all(ws2, 3, ws2.max_row, 1, 4)
    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def generate_stock_excel(shop_id: int) -> bytes:
    pool = await get_pool()
    async with pool.acquire() as conn:
        shop = await conn.fetchrow("SELECT name FROM shops WHERE id=$1", shop_id)
        products = await conn.fetch(
            """SELECT p.name, c.name as cat, p.stock_qty, p.min_stock_qty,
                      p.unit, p.buy_price_som, p.sell_price_som, p.sell_price_usd
               FROM products p LEFT JOIN categories c ON c.id=p.category_id
               WHERE p.shop_id=$1 AND p.is_active=TRUE ORDER BY c.name, p.name""",
            shop_id
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ombor"
    ws["A1"] = f"{shop['name']} — Ombor holati"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    _header_style(ws, 3, ["Mahsulot", "Kategoriya", "Qoldiq", "Min.qoldiq", "Birlik",
                           "Xarid narxi", "Sotuv (so'm)", "Sotuv ($)"])
    total_value = 0
    for p in products:
        val = float(p["stock_qty"]) * float(p["buy_price_som"])
        total_value += val
        row = ws.max_row + 1
        ws.cell(row, 1, p["name"])
        ws.cell(row, 2, p["cat"] or "")
        ws.cell(row, 3, float(p["stock_qty"]))
        ws.cell(row, 4, float(p["min_stock_qty"]))
        ws.cell(row, 5, p["unit"])
        ws.cell(row, 6, float(p["buy_price_som"]))
        ws.cell(row, 7, float(p["sell_price_som"]))
        ws.cell(row, 8, float(p["sell_price_usd"]))
        if float(p["stock_qty"]) <= float(p["min_stock_qty"]):
            for col in range(1, 9):
                ws.cell(row, col).fill = PatternFill("solid", fgColor="FFD700")

    ws.append(["JAMI OMBOR QIYMATI:", "", total_value, "", "", "", "", ""])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    _border_all(ws, 3, ws.max_row, 1, 8)
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
